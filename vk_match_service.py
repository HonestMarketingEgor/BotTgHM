from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

PHONE_ALIASES = [
    "phone",
    "phone_number",
    "telephone",
    "tel",
    "mobile",
    "телефон",
    "номер",
    "номертелефона",
    "мобильный",
]

EMAIL_ALIASES = [
    "email",
    "e-mail",
    "mail",
    "почта",
    "элпочта",
    "login",
]

CAMPAIGN_ID_ALIASES = ["idкампании", "campaignid", "campaign_id", "idcampaign"]
GROUP_ID_ALIASES = ["idгруппы", "adgroupid", "groupid", "group_id"]
AD_ID_ALIASES = ["idобъявления", "adid", "creativeid", "ad_id"]


def _read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype=str)
    if suffix == ".csv":
        for encoding in ("utf-8-sig", "cp1251"):
            try:
                return pd.read_csv(path, dtype=str, sep=None, engine="python", encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Cannot decode CSV file: {path}")
    if suffix == ".pdf":
        return _read_pdf_leads(path)
    raise ValueError(f"Unsupported file format: {path.name}")


def _read_pdf_leads(path: Path) -> pd.DataFrame:
    reader = PdfReader(str(path))
    records = []
    row_pattern = re.compile(
        r"^(?P<registration_id>\d{6,8})"
        r"(?P<registration_status>[^\d]+?)"
        r"(?P<firm_id>\d{5,8})"
        r"(?P<login>[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})"
        r"(?P<rest>.*)$"
    )

    for page_num, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = " ".join(raw_line.split())
            if not line:
                continue
            if line.startswith("-- ") and " of " in line:
                continue
            if line.startswith("RegistrationID") or line.startswith("UtmTermFirstLocalUrlCleaned"):
                continue
            match = row_pattern.match(line)
            if not match:
                continue
            login = match.group("login").strip()
            records.append(
                {
                    "RegistrationID": match.group("registration_id"),
                    "RegistrationStatus": match.group("registration_status").strip() or None,
                    "FirmID": match.group("firm_id"),
                    "Login": login,
                    "raw_line": line,
                    "__pdf_page": page_num,
                    "email": _normalize_email(login),
                    "phone": _normalize_phone(login),
                }
            )

    if not records:
        raise ValueError("No leads parsed from PDF A file.")
    return pd.DataFrame(records)


def _canonical_column_name(name: str) -> str:
    text = str(name).strip().lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def _find_column(
    df: pd.DataFrame,
    aliases: Iterable[str],
    required: bool = False,
    label: str = "column",
) -> Optional[str]:
    canonical_to_original = { _canonical_column_name(col): col for col in df.columns }
    aliases_canonical = [_canonical_column_name(alias) for alias in aliases]

    for alias in aliases_canonical:
        if alias in canonical_to_original:
            return canonical_to_original[alias]
    for canonical_col, original_col in canonical_to_original.items():
        if any(alias in canonical_col for alias in aliases_canonical):
            return original_col
    if required:
        raise ValueError(f"Could not auto-detect required {label}.")
    return None


def _normalize_phone(value: object) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    digits = re.sub(r"\D+", "", str(value))
    if not digits:
        return None
    if len(digits) == 10:
        digits = "7" + digits
    elif len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    return digits


def _normalize_email(value: object) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    email = str(value).strip().lower()
    if not email or "@" not in email:
        return None
    return email


def _is_valid_match_key(value: object) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _validate_result_against_a(df_a_original: pd.DataFrame, result_df: pd.DataFrame) -> None:
    if len(result_df) != len(df_a_original):
        raise ValueError(
            f"Integrity check failed: output row count differs from A ({len(result_df)} vs {len(df_a_original)})."
        )
    a_columns = list(df_a_original.columns)
    missing_columns = [column for column in a_columns if column not in result_df.columns]
    if missing_columns:
        raise ValueError("Integrity check failed: output is missing A columns: " + ", ".join(missing_columns))

    a_compare = df_a_original[a_columns].fillna("").astype(str).reset_index(drop=True)
    result_compare = result_df[a_columns].fillna("").astype(str).reset_index(drop=True)
    row_diff_mask = (a_compare != result_compare).any(axis=1)
    if row_diff_mask.any():
        row_idx = int(row_diff_mask[row_diff_mask].index[0]) + 1
        raise ValueError(f"Integrity check failed: content differs from A at row {row_idx}.")


def _build_top_table(df: pd.DataFrame, source_col: str, out_col: str) -> pd.DataFrame:
    values = df[source_col].dropna().astype(str).str.strip()
    values = values[values != ""]
    counts = values.value_counts(dropna=True)
    if counts.empty:
        return pd.DataFrame(columns=[out_col, "count", "share_percent"])
    top_df = counts.rename_axis(out_col).reset_index(name="count")
    denominator = int(top_df["count"].sum())
    top_df["share_percent"] = (top_df["count"] / denominator * 100).round(2)
    return top_df.sort_values(by=["count", out_col], ascending=[False, True]).reset_index(drop=True)


def _build_top_combinations(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["matched_campaign_id", "matched_group_id", "matched_ad_id"]
    combo_df = df[cols].copy()
    for col in cols:
        combo_df[col] = combo_df[col].fillna("").astype(str).str.strip()
    combo_df = combo_df[
        (combo_df["matched_campaign_id"] != "")
        & (combo_df["matched_group_id"] != "")
        & (combo_df["matched_ad_id"] != "")
    ]
    if combo_df.empty:
        return pd.DataFrame(columns=["campaign_id", "group_id", "ad_id", "count", "share_percent"])

    grouped = (
        combo_df.groupby(cols, dropna=False)
        .size()
        .reset_index(name="count")
        .sort_values(by=["count", *cols], ascending=[False, True, True, True])
        .reset_index(drop=True)
    )
    denominator = int(grouped["count"].sum())
    grouped["share_percent"] = (grouped["count"] / denominator * 100).round(2)
    return grouped.rename(
        columns={
            "matched_campaign_id": "campaign_id",
            "matched_group_id": "group_id",
            "matched_ad_id": "ad_id",
        }
    )


def _write_row(ws, row_idx: int, values: list[object]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def _write_block(ws, start_row: int, title: str, table: pd.DataFrame) -> int:
    _write_row(ws, start_row, [title, "", "", "", ""])
    _write_row(ws, start_row + 1, table.columns.tolist())
    current = start_row + 2
    for row in table.itertuples(index=False):
        _write_row(ws, current, list(row))
        current += 1
    return current + 1


def _append_top_sheet(output_path: Path) -> Dict[str, int]:
    df = pd.read_excel(output_path, dtype=str)
    matched = df["matched"].fillna("").astype(str).str.strip().str.lower().isin({"true", "1", "yes", "y", "да"})
    matched_df = df[matched].copy()

    campaign_top = _build_top_table(matched_df, "matched_campaign_id", "campaign_id")
    group_top = _build_top_table(matched_df, "matched_group_id", "group_id")
    ad_top = _build_top_table(matched_df, "matched_ad_id", "ad_id")
    combo_top = _build_top_combinations(matched_df)

    summary_df = pd.DataFrame(
        [
            ("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("total_rows", len(df)),
            ("matched_rows", len(matched_df)),
            ("combination_rows", int(combo_top["count"].sum()) if not combo_top.empty else 0),
            ("source_file", output_path.name),
        ],
        columns=["metric", "value"],
    )

    wb = load_workbook(output_path)
    if "Top" in wb.sheetnames:
        wb.remove(wb["Top"])
    ws = wb.create_sheet("Top")

    row = 1
    _write_row(ws, row, summary_df.columns.tolist())
    row += 1
    for item in summary_df.itertuples(index=False):
        _write_row(ws, row, list(item))
        row += 1
    row += 1

    row = _write_block(ws, row, "Campaign Top", campaign_top)
    row = _write_block(ws, row, "Group Top", group_top)
    row = _write_block(ws, row, "Ad Top", ad_top)
    _write_block(ws, row, "Top combinations", combo_top)
    wb.save(output_path)

    return {
        "total_rows": len(df),
        "matched_rows": len(matched_df),
        "campaign_top": len(campaign_top),
        "group_top": len(group_top),
        "ad_top": len(ad_top),
        "combination_top": len(combo_top),
    }


def run_vk_match(file_a_path: Path, files_b_paths: list[Path], output_path: Path) -> Dict[str, int]:
    if not file_a_path.exists():
        raise FileNotFoundError(f"File A not found: {file_a_path}")
    if not files_b_paths:
        raise ValueError("No B files provided.")
    for path in files_b_paths:
        if not path.exists():
            raise FileNotFoundError(f"File B not found: {path}")

    df_a = _read_table(file_a_path)
    if df_a.empty:
        raise ValueError("File A is empty.")
    df_a_original = df_a.copy(deep=True)

    b_frames = []
    for path in files_b_paths:
        frame = _read_table(path)
        frame["__source_file"] = path.name
        frame["__source_row"] = range(1, len(frame) + 1)
        b_frames.append(frame)
    df_b = pd.concat(b_frames, ignore_index=True) if b_frames else pd.DataFrame()
    if df_b.empty:
        raise ValueError("Combined B data is empty.")

    a_phone_col = _find_column(df_a, PHONE_ALIASES, required=False, label="A phone column")
    a_email_col = _find_column(df_a, EMAIL_ALIASES, required=False, label="A email column")
    b_phone_col = _find_column(df_b, PHONE_ALIASES, required=False, label="B phone column")
    b_email_col = _find_column(df_b, EMAIL_ALIASES, required=False, label="B email column")
    if not a_phone_col and not a_email_col:
        raise ValueError("No phone or email column detected in file A.")
    if not b_phone_col and not b_email_col:
        raise ValueError("No phone or email column detected in files B.")

    campaign_col = _find_column(df_b, CAMPAIGN_ID_ALIASES, required=True, label="campaign ID")
    group_col = _find_column(df_b, GROUP_ID_ALIASES, required=True, label="group ID")
    ad_col = _find_column(df_b, AD_ID_ALIASES, required=True, label="ad ID")

    df_b["__phone_norm"] = df_b[b_phone_col].map(_normalize_phone) if b_phone_col else None
    df_b["__email_norm"] = df_b[b_email_col].map(_normalize_email) if b_email_col else None
    df_a["__phone_norm"] = df_a[a_phone_col].map(_normalize_phone) if a_phone_col else None
    df_a["__email_norm"] = df_a[a_email_col].map(_normalize_email) if a_email_col else None

    phone_index: Dict[str, pd.Series] = {}
    email_index: Dict[str, pd.Series] = {}
    for _, b_row in df_b.iterrows():
        phone = b_row["__phone_norm"]
        email = b_row["__email_norm"]
        if _is_valid_match_key(phone) and phone not in phone_index:
            phone_index[phone] = b_row
        if _is_valid_match_key(email) and email not in email_index:
            email_index[email] = b_row

    result_rows = []
    matched_count = 0
    for _, a_row in df_a.iterrows():
        phone = a_row["__phone_norm"]
        email = a_row["__email_norm"]
        matched_b_row = None
        match_key = "none"
        if _is_valid_match_key(phone) and phone in phone_index:
            matched_b_row = phone_index[phone]
            match_key = "phone"
        elif _is_valid_match_key(email) and email in email_index:
            matched_b_row = email_index[email]
            match_key = "email"

        row_data = a_row.to_dict()
        row_data["matched"] = bool(matched_b_row is not None)
        row_data["match_key"] = match_key
        row_data["matched_campaign_id"] = matched_b_row[campaign_col] if matched_b_row is not None else None
        row_data["matched_group_id"] = matched_b_row[group_col] if matched_b_row is not None else None
        row_data["matched_ad_id"] = matched_b_row[ad_col] if matched_b_row is not None else None
        row_data["matched_source_file_b"] = matched_b_row["__source_file"] if matched_b_row is not None else None
        row_data["matched_source_row_b"] = matched_b_row["__source_row"] if matched_b_row is not None else None
        if matched_b_row is not None:
            matched_count += 1
        result_rows.append(row_data)

    result_df = pd.DataFrame(result_rows)
    result_df = result_df.drop(columns=[c for c in ["__phone_norm", "__email_norm"] if c in result_df.columns])
    _validate_result_against_a(df_a_original=df_a_original, result_df=result_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    result_df.to_excel(output_path, index=False)
    reloaded = pd.read_excel(output_path, dtype=str)
    if len(reloaded) != len(result_df):
        raise ValueError("Post-write check failed: written XLSX row count mismatch.")

    top_stats = _append_top_sheet(output_path)
    top_stats.update(
        {
            "rows_in_a": len(result_df),
            "matched_rows": matched_count,
            "unmatched_rows": len(result_df) - matched_count,
        }
    )
    return top_stats
